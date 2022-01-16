#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##

# SBK-Charts :  Storage Benchmark Kit - Charts

import argparse
from collections import OrderedDict

import openpyxl
import pandas
import xlsxwriter
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter


class SbkSheets:
    def __init__(self, iFile, oFile):
        self.iFile = iFile
        self.oFile = oFile

    def wb_add_two_sheets(self, wb, r_name, t_name, df):
        header = df.columns.values
        r_ws = wb.add_worksheet(r_name)
        t_ws = wb.add_worksheet(t_name)
        for c, h in enumerate(header):
            r_ws.set_column(c, c, len(h))
            t_ws.set_column(c, c, len(h))
            r_ws.write(0, c, h)
            t_ws.write(0, c, h)
        r_row = 1
        t_row = 1
        for row in df.iterrows():
            if row[1]['Type'] == 'Total':
                for c, h in enumerate(header):
                    col_size = len(str(row[1][h])) + 1
                    if col_size > len(h):
                        t_ws.set_column(c, c, col_size)
                    t_ws.write(t_row, c, row[1][h])
                t_row += 1
            else:
                for c, h in enumerate(header):
                    col_size = len(str(row[1][h])) + 1
                    if col_size > len(h):
                        r_ws.set_column(c, c, col_size)
                    r_ws.write(r_row, c, row[1][h])
                r_row += 1

    def create_sheets(self):
        df = pandas.read_csv(self.iFile)
        wb = xlsxwriter.Workbook(self.oFile)
        self.wb_add_two_sheets(wb, "R1", "T1", df)
        wb.close()
        print("xlsx file %s created" % self.oFile)


class SbkCharts:
    def __init__(self, file):
        self.file = file

    def get_columns_from_worksheet(self, ws):
        return {
            cell.value: {
                'letter': get_column_letter(cell.column),
                'number': cell.column
            } for cell in ws[1] if cell.value
        }

    def get_latency_columns(self, ws):
        column_names = self.get_columns_from_worksheet(ws)
        ret = OrderedDict()
        ret['AvgLatency'] = column_names['AvgLatency']
        ret['MaxLatency'] = column_names['MaxLatency']
        for key in column_names.keys():
            if key.startswith("Percentile_"):
                ret[key] = column_names[key]
        return ret

    def get_time_unit(self, ws):
        colnames = self.get_columns_from_worksheet(ws)
        return ws.cell(row=2, column=colnames['LatencyTimeUnit']['number']).value


    def create_latency_compare_graphs(self, wb, ws, time_unit, prefix):
        latencies = self.get_latency_columns(ws)
        charts = [LineChart(), LineChart(), LineChart(), LineChart()]

        for ch in charts:
            # set the title of the chart
            ch.title = " Percentile Variations"
            # set the title of the x-axis
            ch.x_axis.title = " Intervals "
            # set the title of the y-axis
            ch.y_axis.title = " Latency Time in " + time_unit
            ch.height = 25
            ch.width = 50

        sheets = [wb.create_sheet("Latencies-1"), wb.create_sheet("Latencies-2"),
                 wb.create_sheet("Latencies-3"), wb.create_sheet("Latencies-4")]

        groups = [
            ["Percentile_10", "Percentile_20", "Percentile_25", "Percentile_30", "Percentile_40", "Percentile_50"],
            ["Percentile_50", "AvgLatency"],
            ["Percentile_50", "Percentile_60", "Percentile_70", "Percentile_75", "Percentile_80", "Percentile_90"],
            ["Percentile_92.5", "Percentile_95", "Percentile_97.5", "Percentile_99",
             "Percentile_99.25", "Percentile_99.5", "Percentile_99.75", "Percentile_99.9",
             "Percentile_99.95", "Percentile_99.99"]]

        for x in latencies:
            data_series = Series(Reference(ws, min_col=latencies[x]['number'], min_row=2,
                                           max_col=latencies[x]['number'], max_row=ws.max_row),
                                 title=prefix+"-"+x)
            for i, g in enumerate(groups):
                if x in g:
                    charts[i].append(data_series)
        for i, ch in enumerate(charts):
            sheets[i].add_chart(ch)


    def create_latency_graphs(self, wb, ws, time_unit, prefix):
        latencies = self.get_latency_columns(ws)
        for x in latencies:
            data_series = Series(Reference(ws, min_col=latencies[x]['number'], min_row=2,
                                           max_col=latencies[x]['number'], max_row=ws.max_row),
                                 title=prefix+"-"+x)
            chart = LineChart()
            # adding data
            chart.append(data_series)
            # set the title of the chart
            chart.title = x + " Variations"
            # set the title of the x-axis
            chart.x_axis.title = " Intervals "
            # set the title of the y-axis
            chart.y_axis.title = " Latency Time in " + time_unit
            chart.height = 25
            chart.width = 45
            # add chart to the sheet
            sheet = wb.create_sheet(x)
            sheet.add_chart(chart)


    def create_throughput_MB_graph(self, wb, ws, prefix):
        cols = self.get_columns_from_worksheet(ws)
        data_series = Series(Reference(ws, min_col=cols["MB/Sec"]['number'], min_row=2,
                                       max_col=cols["MB/Sec"]['number'], max_row=ws.max_row),
                             title=prefix+"-MB/Sec")
        chart = LineChart()
        # adding data
        chart.append(data_series)
        # set the title of the chart
        chart.title = " Throughput Variations in Mega Bytes / Seconds"
        # set the title of the x-axis
        chart.x_axis.title = "Intervals"
        # set the title of the y-axis
        chart.y_axis.title = "Throughput in MB/Sec"
        chart.height = 20
        chart.width = 40
        # add chart to the sheet
        sheet = wb.create_sheet("MB_Sec")
        sheet.add_chart(chart)

    def create_throughput_records_graph(self, wb, ws, prefix):
        cols = self.get_columns_from_worksheet(ws)
        data_series = Series(Reference(ws, min_col=cols["Records/Sec"]['number'], min_row=2,
                                       max_col=cols["Records/Sec"]['number'], max_row=ws.max_row),
                             title=prefix+"-Records/Sec")
        chart = LineChart()
        # adding data
        chart.append(data_series)
        # set the title of the chart
        chart.title = " Throughput Variations in Records / Second"
        # set the title of the x-axis
        chart.x_axis.title = " Intervals "
        # set the title of the y-axis
        chart.y_axis.title = "Throughput in Records/Sec"
        chart.height = 20
        chart.width = 40
        # add chart to the sheet
        sheet = wb.create_sheet("Records_Sec")
        sheet.add_chart(chart)

    def create_graphs(self):
        wb = openpyxl.load_workbook(self.file)
        r_name = "R1"
        t_name = "T1"
        ws1 = wb[r_name]
        ws2 = wb[t_name]
        """
        for row in ws1.iter_rows():
            lt = []
            for cell in row:
                lt.append(cell.value)
            print(lt)
        """
        self.create_throughput_MB_graph(wb, ws1, r_name)
        self.create_throughput_records_graph(wb, ws1, r_name)
        self.create_latency_compare_graphs(wb, ws1, self.get_time_unit(ws1), r_name)
        self.create_latency_graphs(wb, ws1, self.get_time_unit(ws1), r_name)
        wb.save(self.file)


def main():
    parser = argparse.ArgumentParser(description='sbk charts')
    parser.add_argument('-i', '--ifile', help='Input CSV file', required=True)
    parser.add_argument('-o', '--ofile', help='Output xlsx file', default="out.xlsx")
    args = parser.parse_args()
    print('Input file is ', args.ifile)
    print('Output file is ', args.ofile)
    sheets = SbkSheets(args.ifile, args.ofile)
    sheets.create_sheets()
    charts = SbkCharts(args.ofile)
    charts.create_graphs()


if __name__ == "__main__":
    main()
